#!/usr/bin/env python
"""
Project:    CLIPRT - Client Information Parsing and Reporting Tool.
@author:    mhodges
Copyright   2020 Michael Hodges
"""
import openpyxl
import os.path

from cliprt.classes.identifier_registry import IdentifierRegistry
from cliprt.classes.client_registry import ClientRegistry
from cliprt.classes.content_worksheet import ContentWorksheet
from cliprt.classes.data_element_dictionary_processor import DataElementDictionaryProcessor
from cliprt.classes.destination_worksheets_registry import DestinationWorksheetsRegistry
from cliprt.classes.message_registry import MessageRegistry

class ClientInformationWorkbook:
    """
    The client information workbook orchestrates the various activities
    performed across the entire set of worksheets, including worksheet 
    creation as needed.  It is the primary backend to the cliprt command
    line user interface.
    """
    # Dependencies
    error = MessageRegistry()

    # Internal worksheets names.
    DED_WS_NAME = 'Data Elements'
    INTERNAL_WS_NAMES = [DED_WS_NAME]

    # Reports worksheets name prefix.
    DEST_WS_NAME_PREFIX = 'cliprt_report_for_'

    def __init__(self, wb_filename):
        """
        Ensure that the workbook exists.  Set everything up for 
        processing the workbook.
        """
        if not os.path.exists(wb_filename):
            # Fatal error
            raise Exception(self.error.msg(1000).format(wb_filename))

        # Class attributes.
        self.ded_processor = None
        self.ded_ws = None
        self.dest_ws_reg = DestinationWorksheetsRegistry()
        self.client_reg = ClientRegistry(self.dest_ws_reg)
        self.content_ws_names = []
        self.identifier_reg = IdentifierRegistry()
        self.wb = openpyxl.load_workbook(filename=wb_filename)
        self.wb_filename = wb_filename

        # The workbook may or may not have a DED worksheet when it is 
        # initially accessed.
        self.init_ded_processor()

    def create_client_reports(self, progress_reporting_is_disabled = False, save_wb = True):
        """
        Utilize and process the various worksheets in the workbook in
        order to create the destination report worksheets.
        """
        # Create the DED.  This will also create a destination worksheet
        # registry.
        self.ded_processor.hydrate_ded()

        # Create or reset the destination worksheets in preparation for
        # the next round of reports.
        self.ded_processor.dest_ws_reg.prep_worksheets()
        
        # Create the list of client data content worksheets.
        self.create_content_ws_names_list()

        # Process contents of each client data worksheet.
        #ContentWorksheet()
        for ws_name in self.content_ws_names:
            ContentWorksheet(
                self.wb,
                ws_name,
                self.ded_processor, 
                self.client_reg,
                self.identifier_reg,
                self.dest_ws_reg
            ).client_report(progress_reporting_is_disabled)
        
        # Save the client report worksheets.
        if save_wb:
            self.wb.save(self.wb_filename)
        
    def create_content_ws_names_list(self):
        """
        Create the list of data content worksheet names.  This list
        will be looped through to produce the client reports.
        """
        for ws_name in self.wb.sheetnames:
            if ws_name in self.INTERNAL_WS_NAMES:
                # Ignore the internal worksheets.
                continue
            if ws_name in self.dest_ws_reg.dest_ws_names:
                # Ignore any existing destination worksheets.
                continue
            # Save the data content worksheet name.
            self.content_ws_names.append(ws_name)

    def create_de_names_list(self):
        """
        Read through all of the content worksheets and create a unique,
        sorted list of the data element names.  This list will be used
        to create a new data element dictionary.
        """
        de_names = []
        if len(self.content_ws_names) == 0:
            self.create_content_ws_names_list()
        for ws_name in self.content_ws_names:
            ws = self.wb[ws_name]
            for cell in ws[ws.min_row]:
                if cell.value == None:
                    continue
                de_name = cell.value.lower().strip().title()
                if len(de_name) >= 2 and not de_name in de_names:
                    de_names.append(de_name)
        de_names.sort()
        return de_names

    def create_ded_worksheet(self, save_wb = True):
        """
        Create a fresh DED worksheet for configuration.
        """
        if self.DED_WS_NAME in self.wb:
            # The DED worksheet already exists.
            return
        
        # Create the DED worksheet.
        self.ded_ws = self.wb.create_sheet(title=self.DED_WS_NAME, index=0)
        de_names = self.create_de_names_list()

        # Now that we have a DED worksheet we can instantiate the DED
        # processor and the destination worksheet registry.
        self.init_ded_processor()
        self.ded_processor.preconfig_ded_worksheet(de_names)

        # Save the DED worksheet, unless running unit tests, for example.
        if save_wb:
            self.wb.save(self.wb_filename)

    def ded_is_configured(self):
        """
        Check to see if the DED has been configured.
        """
        if self.ded_processor == None:
            # The DED worksheet isn't available or the DED processor
            # has not been instantiated yet.
            return False
        return True if self.ded_processor.ded_is_configured() else False

    def init_ded_processor(self):
        """
        If the DED worksheet is available initialize the DED processor.
        """
        if not self.has_a_ded_ws():
            return False
        self.ded_ws = self.wb[self.DED_WS_NAME]
        self.ded_processor = DataElementDictionaryProcessor(self.wb, self.ded_ws, self.dest_ws_reg)
        return True

    def has_a_ded_ws(self):
        """
        Check to see if the client information workbook has a 
        DED worksheet.
        """
        return True if self.DED_WS_NAME in self.wb.sheetnames else False
    
    def print_ded_report(self):
        """
        Print the data element dictionary contents.  Useful for 
        reviewing.
        """
        self.ded_processor.print_report()