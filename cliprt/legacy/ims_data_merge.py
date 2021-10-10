#!/usr/bin/env python
"""
Project:  CLIPRT - Client Information Parsing and Reporting Tool.
Version:  0.1.1
@author:  mhodges
Copyright 2020 Michael Hodges

Background:
    Version 0.1.x was implemented as procedural code while I focused on 
    learning the syntax and the pythonista perspective.  Version 0.2+ was 
    always intended to be object oriented (OO).

Relevations:
    The structure of the procedural code mapped pretty easily to OO code. 
    The OO code is much easier to optimize, refacter and debug, and a lot 
    more fun to implement overall as each object takes on its own
    personality.
"""
import os
import openpyxl
import ims_util

# ----------------
# Assign constants
# ----------------

# Abbreviations:
# - COL is a worksheet column
# - DE is a data element, e.g.:  "first name"
# - DED is the data element dictionary
# - DEST is the destination worksheet
# - WB is a workbook 
# - WS is a worksheet

# For now this is the workbook.  Later any workbook will be supported.
IMS_WORKBOOK = '/Users/mhodges/VisualStudioCodeProjects/cliprt/cliprt/resources/sample.xlsx'
DEBUG_WORKBOOK = '/Users/mhodges/VisualStudioCodeProjects/cliprt/cliprt/resources/debug_sample.xlsx'

# Debugging.
DEBUG_REPORTING_ONLY = True

# Internal worksheets names.
DE_WS_NAME = 'Data Elements'
ID_MATCHING_LOG_WS_NAME = 'id_matching_log'
INTERNAL_WS_NAMES = [ID_MATCHING_LOG_WS_NAME, DE_WS_NAME]

# Reports worksheets name prefix...
DEST_WS_NAME_PREFIX = 'comm_report_for_'

# Required DED column headings.
DED_COL_HEADINGS = ['Data Element', 'Dest WS', 'Dest Element', 'DE Format']
DED_COL_IDX = 0
DEST_WS_COL_IDX = 1
DEST_DED_COL_IDX = 2
DEST_DE_FORMAT_COL_IDX = 3

# Valid data element formats.  A fatal error will occur if invalid data is encountered.
VALID_DE_FORMATS = ['identifier', 'fragment', 'date', 'name', 'phone']

# Threshold for determining that we have an identity match.
IDENTITY_MATCH_THRESHOLD = 2 

# Id matching log column headers.
ID_MATCHING_LOG_HEADERS = ['Identifiers Matched', 'Identifiers Not Matched', 'Outcome', 'Content Worksheet', 'Destination Worksheet(s)']
LOG_ID_MATCHED = 0
LOG_ID_NOT_MATCHED = 1
LOG_OUTCOME = 2
LOG_CONTENT_WS = 3
LOG_DEST_WS = 4

# ---------------------------------------------
# Open the workbook and checck the requirements
# ---------------------------------------------
wb = openpyxl.load_workbook(filename=IMS_WORKBOOK)

# Ensure that the Data Elements worksheet is available.  Error out if not.
if not DE_WS_NAME in wb.sheetnames: 
    # Fatal error
    raise Exception('WS Error: required worksheet {} not found.\nUnable to continue.  See example worksheet {} for comparison.'.format(DE_WS_NAME, IMS_WORKBOOK))

# Get the Data Elements worksheet.
ws = wb[DE_WS_NAME]

# Get the column headings. Retain also the column index.
col_headings = {}
for cell in ws[ws.min_row]:
    col_headings[cell.value] = cell.col_idx

# Ensure that all of the required named columns are available.  Error out if not.
for ded_col_heading in DED_COL_HEADINGS:
    if ded_col_heading not in col_headings:
        # Fatal error
        raise Exception('WS Error: required column heading {} not found in worksheet {}.\nUnable to continue.  See example worksheet {} for comparison.'.format(ded_col_heading, DE_WS_NAME, IMS_WORKBOOK))

# --------------
# Create the DED
# --------------
# Set up the metadata for each data element in the data element dictionary and for each destination worksheet.

# Metadata initializations.
ded = {}                # Data element dictionary
dest_ws_names = []      # Destination worksheet names list
valid_de_list = []      # Valid data element names list
dest_ws_by_ind = {}     # Destinations worksheets by worksheet indicator
de_fragments_list = {}  # Data elements fragments by data element name list.

# Initialize the identity matching dictionary, where:
# > identity['last_id'] = 1000
# > identity['id'][1000] = {'dest_ws': {ims': {'row': [n]}, 'fb': {'row': [n]}}, 'identifier': {'de_namme': []}}
# > identity['value'][val] = {'de_name': de_name, 'id_nos': set(n)}
identity = {'id': {}, 'last_id': 999, 'value': {}}

# Get the DED column of data element names.
ws_columns = ws.iter_cols(min_col=col_headings[DED_COL_HEADINGS[DED_COL_IDX]], max_col=col_headings[DED_COL_HEADINGS[DED_COL_IDX]], min_row=ws.min_row+1)
de_column = list(ws_columns)[0]

# From each cell get the data element name and associated information.
for de_cell in de_column:

    # Skip rows that are missing the data element name.
    if de_cell.value == None:
        continue
        
    # Ensure values used for comparisons are shifted to lowercase to reduce sensitivity to typos in the DED.
    de_name = de_cell.value.lower()
    dest_element = ims_util.lowercase(ws.cell(row=de_cell.row, column=col_headings[DED_COL_HEADINGS[DEST_DED_COL_IDX]]).value)

    # Skip elements that have neither a defined destination worksheet nor are mapped to another data element.
    dest_ws_indicators = ims_util.lowercase(ws.cell(row=de_cell.row, column=col_headings[DED_COL_HEADINGS[DEST_WS_COL_IDX]]).value)
    if dest_ws_indicators == None and dest_element == None:
        continue
    if not dest_ws_indicators == None and not dest_element == None:
        # Fatal error
        raise Exception('DED Error: specify either a "Dest WS" or a "Dest Element" for {} in worksheet {}, but not both.\nUnable to continue.  See example worksheet {} for comparison.'.format(de_name, DE_WS_NAME, IMS_WORKBOOK))

    # -----------------------------
    # Initialize the new DED entry.
    # -----------------------------
    if not de_name in ded:
        ded[de_name] = {}

    # Add the data element name and its information to the data element dictionary.
    if not dest_element == None:
        ded[de_name]['dest_de_name'] = dest_element

    # The dest_format is overloaded with the "identifier" and "fragment=n" designaters.
    dest_format_str = ims_util.lowercase(ws.cell(row=de_cell.row, column=col_headings[DED_COL_HEADINGS[DEST_DE_FORMAT_COL_IDX]]).value)
    if not dest_format_str == None:

        # Initialize the clean list of destination formats.
        dest_format_list = []

        # Preprocess the raw destination format list and parse "fragment" entries to get the fragment index. 
        for dest_format in ims_util.make_list(dest_format_str):

            # Determine if the destination format needs to be parsed.
            if '=' in dest_format:

                # Parse the destination format from the assigned index.
                dest_format_part = dest_format.split('=',1)

                # Parsed, this is the clean destination format.
                dest_format = dest_format_part[0]

                # Currently the assigned index is only relevant for a "fragment", the rest are ignored.
                if dest_format == 'fragment':
                    if not dest_format_part[1].isdigit():
                        # Fatal error
                        raise Exception('DED Error: Destination format error for {} in worksheet {}. Fragment="n" where "n" is an integer is required.\nUnable to continue.  See example worksheet {} for comparison.'.format(de_name, DE_WS_NAME, IMS_WORKBOOK))

                    # Save the fragment index for later.
                    de_fragments_list[de_name] = int(dest_format_part[1])

            # Add the clean destination format to the list.
            dest_format_list.append(dest_format)

            if 'identifier' in dest_format_list and 'fragment' in dest_format_list:
                # Fatal error, fragments can only reference an identifier.
                raise Exception('DED Error: invalid destination format "{}" specified for "{}". A fragment can only reference an identifier.\nUnable to continue until you remove "identifier".'.format(dest_format, de_name))

        # Process the clean list of destination formats.
        for dest_format in dest_format_list:

            # Validate the destintion format designater.
            if not dest_format in VALID_DE_FORMATS:
                # Fatal error, invalid destination format
                raise Exception('DED Error: invalid destination format "{}" specified for "{}".\nUnable to continue.  Valid values: {}.'.format(dest_format, de_name, VALID_DE_FORMATS))

            # Check for the overload identifier and fragment designaters.
            if dest_format == 'identifier':
                ded[de_name]['is_identifier'] = True

            elif dest_format == 'fragment':
                if de_name in de_fragments_list:
                    ded[de_name]['fragment_idx'] = de_fragments_list[de_name]
                else:
                    # Fatal error, invalid destination format
                    raise Exception('DED Error: invalid destination format "{}" specified for "{}".  fragment="n" expected.\nUnable to continue.  Valid values: {}.'.format(dest_format, de_name))
                
            else:
                # Save the destintion format designater to the DED.
                ded[de_name]['dest_format'] = dest_format

    # Create the list for later validation that all of the destinations data elements are in the DED.
    if not dest_element == None and not dest_element in valid_de_list:
        valid_de_list.append(dest_element)

    # Data elements mapped to a destination element don't contain useful destination worksheet information.
    if not dest_element == None:
        continue

    # For each data element more than one destination worksheet may be specified in the DED worksheet.
    # Process each indicator's specified worksheet.
    for ws_dest_ind in ims_util.make_list(dest_ws_indicators):

        # Autodetect new destination indicators.
        if not ws_dest_ind in dest_ws_by_ind:

            # New destination worksheet indicator detected.  Initializing.
            dest_ws_by_ind[ws_dest_ind] = {'ws': {}, 'last_col': 0, 'last_row': 0, 'dest_de_list': {}}

            # Update the list of destination worksheet names.  
            # Later will need to skip these while processing the data content worksheets.
            if not DEST_WS_NAME_PREFIX + ws_dest_ind in dest_ws_names:
                dest_ws_names.append(DEST_WS_NAME_PREFIX + ws_dest_ind)

        # Each column heading is assigned the next available column.
        # Increment the index of the last data element column of the specified destination worksheet.
        dest_ws_by_ind[ws_dest_ind]['last_col'] += 1

        # Save the data elements target worsheet indicator, column index, and mapped name for reference.
        actual_de_element = dest_element if not dest_element == None else de_name
        if not 'dest_ws' in ded[de_name]:
            ded[de_name]['dest_ws'] = {}
        ded[de_name]['dest_ws'][ws_dest_ind] = {'column': dest_ws_by_ind[ws_dest_ind]['last_col']}
        # This will be used to create the column headings for each of the dstination worksheets.
        dest_ws_by_ind[ws_dest_ind]['dest_de_list'][de_name] = dest_ws_by_ind[ws_dest_ind]['last_col']

# Validate that all destination data elements are in the DED.
for dest_element in valid_de_list:
    if not dest_element in ded:
        # Fatal error
        raise Exception('DED Error: see column "Dest Element" for invalid entry "{}".  It must reference a entry in the colum "Data Element".\nUnable to continue until you correct the worksheet.'.format(dest_element, DE_WS_NAME))

# -----------------------------------
# Isolate the data content worksheets
# -----------------------------------

data_content_ws_names = []

# Create the list of data content worksheet names.
for ws_name in wb.sheetnames:

    # Ignore the internal worksheets.
    if ws_name in INTERNAL_WS_NAMES:
        continue
    # Ignore any existing destination worksheets.
    if ws_name in dest_ws_names:
        continue
    # Save the data content worksheet name.
    data_content_ws_names.append(ws_name)

# ----------------------------------
# Prepare the destination worksheets
# ----------------------------------

# Prepare a fresh set of destination worksheets.
for ws_dest_ind in dest_ws_by_ind:

    ws = ims_util.init_ws(wb, DEST_WS_NAME_PREFIX + ws_dest_ind)
    last_row = 1

    # Add the column headings to the destination worksheet.
    for col_title in dest_ws_by_ind[ws_dest_ind]['dest_de_list']:
        col_idx = dest_ws_by_ind[ws_dest_ind]['dest_de_list'][col_title]
        ws.cell(last_row, col_idx, value=col_title)

    # Initialize for later access to the destintation worksheet.
    dest_ws_by_ind[ws_dest_ind]['ws'] = ws
    dest_ws_by_ind[ws_dest_ind]['last_row'] = last_row

# =====================================================================
# Progress report
print()
print("-------------------------------------------")
print("Success: data elemement dictionary created.")
print('> ded.items()')
print("-------------------------------------------")
for key, details in ded.items():
    print('{} : {}'.format(key, details))
print()
print("--------------------------------------------------")
print("Initialized: destination worksheet data strutures.")
print('> dest_ws_by_ind.items()')
print("--------------------------------------------------")
for key, details in dest_ws_by_ind.items():
    print('{} : {}'.format(key, details))
print()
print("-----------------------------------------------")
print("Initialized: identity matching data structures.")
print('> identity.items()')
print("-----------------------------------------------")
for key, details in identity.items():
    print('{} : {}'.format(key, details))
print()
# =====================================================================

# ------------------------------------
# Initialize the identity matching log
# ------------------------------------

# Initializing just in time.
ims_util.init_ws(wb, ID_MATCHING_LOG_WS_NAME)
ws = wb[ID_MATCHING_LOG_WS_NAME]

# If the column headings have not been created, create them.
if ws.max_column < 2:

    # Populate the column headers.
    col_idx = 1
    for col_heading in ID_MATCHING_LOG_HEADERS:
        ws.cell(ws.min_row, col_idx, value=col_heading)
        col_idx+=1

# ------------------------------
# Process the content worksheets
# ------------------------------

# Process each data content worksheet.
for ws_name in data_content_ws_names:

    # -----------------------------------------------
    # Build the destination data element ETL mappings
    # -----------------------------------------------
    # These ETL mappings will provde exact instructions for making each 
    # content data element to each of the destination worksheets.

    # Open the first or next content worksheet.
    ws = wb[ws_name]

    # Initialize where:
    # > ws_column_etl[col_idx] = {'dest_ws': {dest_ws: dest_col_idx}, 'dest_format': dest_format}
    ws_column_etl = {}

    # Initialize lists.
    ws_identifier_cols = []
    ws_content_cols = []
    dest_ws_ind_list = []
    ws_de_names = []

    # Initialize where:
    # > dest_de_assembly[dest_de_name] = {de_fragment_name: de_frag_col_idx}
    dest_de_fragments = {}

    # First get a list of all of the data element names for this content worksheet.  
    # We will need this list to process fragments later.
    ws_top_rows = ws.iter_cols(min_row=ws.min_row, max_row=ws.min_row)
    ws_top_row = list(ws_top_rows)[0]

    # ------------------------------
    # Build content workshet DE list
    # ------------------------------
    # Preread the top row in order to build the list of the data element names in this worksheet.
    for ws_cell in ws_top_row:

        # Skip empty columns.
        if ws_cell.value == None:
            continue

        # Perform a DED lookup and skip this column if this content data element is not in the DED.
        de_name = ws_cell.value.lower()
        if not de_name in ded:
            continue;

        # Build the list of the content data element names (column names).
        ws_de_names.append(de_name)

    # -------------------------------
    # Build content worksheet ETL map
    # -------------------------------
    # Read the top row and build the ETL mappings to the destination worksheets.
    # Also flag the identity and fragment columns for special processing.
    for ws_cell in ws_top_row:

        # Skip empty columns.
        if ws_cell.value == None:
            continue

        # Perform a DED lookup and skip this column if this content data element is not in the DED.
        ws_de_name = ws_cell.value.lower()

        # todo: log the skips
        if not ws_de_name in ded:
            continue;

        # Initialize.
        is_content = True
        is_fragment = False
        fragment_idx = int()

        # Copy the worksheet data element name in case it gets remapped to a destination
        # worksheet name.
        de_name = ws_de_name
        
        # Identify the content worksheet data elements, resolve data element mapping requirements.
        if 'dest_de_name' in ded[de_name]:

            # This content data element is mapped to another data elememnt, which must specify a destination worksheet.
            # Before mapping the name, also check to see if it is a fragment.
            if 'fragment_idx' in ded[de_name]:
                is_content = False
                is_fragment = True
                fragment_idx = ded[de_name]['fragment_idx']

            # Repoint to the provided destination data element.  This is a remapping: a->b rather than a->a.
            de_name = ded[de_name]['dest_de_name']
            if not 'dest_ws' in ded[de_name]:
                # Fatal error: there must be a destination worksheet.
                raise Exception('DED Error, data element "{}" specifies an invalid destination: "{}".\nUnable to continue until you correct the DED.'.format(ws_cell.value, de_name))

        # ETL mapping: map this data content column to the target destination worksheet column.
        ws_column_etl[ws_cell.column] = {'dest_ws': {}}
        for dest_ws in ded[de_name]['dest_ws']:
            ws_column_etl[ws_cell.column]['dest_ws'][dest_ws] = ded[de_name]['dest_ws'][dest_ws]['column']
            if 'dest_format' in ded[de_name]:
                ws_column_etl[ws_cell.column]['dest_format'] =  ded[de_name]['dest_format']
            if not dest_ws in dest_ws_ind_list:
                dest_ws_ind_list.append(dest_ws)

        # Include destination fragment information in the mapping if needed and add the
        # new column to hold the assembled fragments if needed.
        if is_fragment:

            # Map the composition of the assembled data element name for later assembly.
            if not de_name in dest_de_fragments:
                dest_de_fragments[de_name] = {'frag_de_name': {}}
            dest_de_fragments[de_name]['frag_de_name'][ws_de_name] = {'frag_idx': fragment_idx, 'col_idx': ws_cell.column}

            # If not found create the new column heading.
            if not de_name in ws_de_names:
                # Add a new column to the worksheet.
                ws_de_names.append(de_name)

        # Flag the column as an identifier if it is one.
        # Note that fragmented identifiers are not identifiers until the fragments are assembled.
        elif 'is_identifier' in ded[de_name]:
            ws_column_etl[ws_cell.column]['identifier'] = de_name

            # Add the column index to the list of identifier columns.
            ws_identifier_cols.append(ws_cell.column)

            # Identifiers are actually content too, but they get processed before simple content is processed.
            is_content = False

        # Add the column index to the list of content columns.
        if is_content:
            ws_content_cols.append(ws_cell.column)

    # =================================================================
    # Progress report
    print('--------')
    print('Success: workbook "{}" DE map created.'.format(ws_name))
    print('> ws_column_etl.items{}')
    for key, details in ws_column_etl.items():
        print('{} : {}'.format(key, details))
    print("Dest WSs    > dest_ws_ind_list   : {}".format(dest_ws_ind_list))
    print("DE Names    > ws_de_names        : {}".format(ws_de_names))
    print("> dest_de_fragments{}")
    for key, details in dest_de_fragments.items():
        print('{} : {}'.format(key, details))
    print("Identifiers > ws_identifier_cols : {}".format(ws_identifier_cols))
    print("Content     > ws_content_cols    : {}".format(ws_content_cols))
    # =================================================================
    if DEBUG_REPORTING_ONLY:
        continue
    # =================================================================

    # -------------------------------
    # Skip worksheets with no content
    # -------------------------------
    if len(ws_content_cols) == 0:
        continue

    # -----------------------------------
    # Populate the destination worksheets
    # -----------------------------------
    # Process the content worksheet rows. 
    # Steps:
    # 1. Process the destination data element fragments.
    # 2. Process the identifiers.
    #    - a perfect match is found if all identifiers match.
    #    - a likely match is found if only one identifiers doesn't match.
    #    - determine if data is being saved to an exiting row or to a new row.
    # 3. Save the data.

    # Initialize.
    row_idx = ws.min_row 
    while row_idx < ws.max_row:
        row_idx+=1

        # ---------------------------------
        # Assemble fragmented data elements
        # ---------------------------------

        # Initialize where:
        # > assembled_de_values[de_type][de_name] = {'dest_ws': {}, 'value': ims_util.list_to_str(de_values_list)}        
        assembled_de_values = {}
        if len(dest_de_fragments) > 0:

            # Initialize the fragment lists by pre-sizing them to hold all fragments.
            # Fragments are assembled based on the destination data element name.
            fragment_dest_names = {}
            for dest_de_name, de_details in dest_de_fragments.items():
                fragment_dest_names[dest_de_name] = ims_util.init_list(len(de_details))

            # Assemble the fragmented data element.
            de_fragments = {}
            for dest_de_name, de_details in dest_de_fragments.items():
                de_fragments[dest_de_name] = ims_util.init_list(len(de_details['frag_de_name']))
                for de_name, frag_details in de_details['frag_de_name'].items():
                    col_idx = frag_details['col_idx']
                    frag_value = ws.cell(row_idx, col_idx).value
                    frag_idx = frag_details['frag_idx'] - 1
                    de_fragments[dest_de_name][frag_idx] = frag_value

            # Create a separate ETL mapping for the assembled fragments.
            for de_name, de_values_list in de_fragments.items():

                de_type = 'content' if 'is_identifier' not in ded[de_name] else 'identifier' 

                # Keep track of the assemble identifiers and content for later storage.
                if not de_type in assembled_de_values:
                    assembled_de_values[de_type] = {}

                # Assemble the fragments into a single value string.
                de_value = ims_util.list_to_str(de_values_list)

                # Assembled identifiers need special handling.
                if de_type == 'is_identifier':
                    # Identity data is stored in lowercase.
                    de_value = ims_util.make_searchable(de_value)

                # Keep track of the assemble identifiers and content for later storage.
                assembled_de_values[de_type][de_name] = {'dest_ws': {}, 'value': de_value}
                assembled_de_values[de_type][de_name]['dest_ws'] = ded[de_name]['dest_ws']

        # -----------------------
        # Process the identifiers
        # -----------------------
        # o Attempt to match and merge identities when it is determined that we have multiple
        #   records belonging to the same person.
        # o Determine which row to use for the each of the destination worksheets.

        # The destination row indices will be determined based on whether the data matches an
        # exiting identity or if the row represents a brand new identity.

        # Key assumption(s):
        # 1. If at least two unique identifier types match, we have a probable identity match.

        # TODO: introduce a dest wordsheet column that reports that we have performed an identity merge.
        # TODO: introduce yet another worksheet that journals successful and rejected identity merges.

        # Initialize list where, for either 'matched' or 'unmatched':
        # > identifier_matching['matched'][de_name] = {'value': val}
        identifier_matching = {'matched': {}, 'unmatched': {}}
        identifier_types = []
        identity_idno = int()

        # Check for and process the identifiers.
        for col_idx in ws_identifier_cols:
            de_value = ws.cell(row_idx, col_idx).value
            if de_value == None:
                continue
            de_name = ws_column_etl[col_idx]['identifier']
            de_value = ims_util.make_searchable(de_value)
            ims_util.save_identifier(identity, identifier_matching, de_name, de_value, identifier_types)

        # Check for and process any assembled identifiers.
        if 'identifier' in assembled_de_values:
            for de_name, de_details in assembled_de_values['identifier'].items():
                de_value = ims_util.make_searchable(de_details['value'])
                ims_util.save_identifier(identity, identifier_matching, de_name, de_value, identifier_types)

        # Determine if there is a significant number of matches against an existing identity.
        if len(identifier_types) < IDENTITY_MATCH_THRESHOLD:

            # ------------------------
            # Set up a new identity id
            # ------------------------

            # Set up the new identifier number.
            identity_idno = identity['last_id'] + 1
            identity['id'][identity_idno] = {}                

            # Update the last row values.
            for dest_ws_ind, dest_ws_details in dest_ws_by_ind.items():

                # Reference the next available destination worksheet row.
                dest_row_idx = dest_ws_details['last_row'] + 1

                # Update tne 'last row' reference for each destination worksheet.
                dest_ws_details['last_row'] = dest_row_idx
                if not 'dest_ws' in identity['id'][identity_idno]:
                    identity['id'][identity_idno] = {'dest_ws': {dest_ws_ind: {}}}
                identity['id'][identity_idno]['dest_ws'][dest_ws_ind] = {'row': dest_row_idx}

            for de_name, de_detail in identifier_matching['matched'].items():
                de_value = de_detail['value']
                idno_set = {identity_idno}
                identity['value'][de_value] = {'de_name': de_name, 'idno_set': idno_set}

            identity['id'][identity_idno]['identifiers'] = identifier_matching['matched']
            identity['last_id'] = identity_idno

        else:

            # ------------------------------
            # Match the existing identity id
            # ------------------------------

            idno_sets = []
            for de_name, de_details in identifier_matching['matched'].items():
                de_value = de_details['value']
                idno_set = identity['value'][de_value]['idno_set']
                idno_sets.append(idno_set)

            # Find the best match.
            identity_idno = ims_util.id_number_matcher(idno_sets)

        # Now that the identity number is known, update its identifiers if needed.
        idno_set = {identity_idno}
        for de_name, de_details in identifier_matching['matched'].items():
            de_value = de_details['value']
            saved_idno_set = identity['value'][de_value]['idno_set']
            if not idno_set.issubset(saved_idno_set):
                saved_idno_set.add(identity_idno)

        # Update the unmatched data elements too.
        for de_name, de_details in identifier_matching['unmatched'].items():
            de_value = de_details['value']
            saved_idno_set = identity['value'][de_value]['idno_set']
            if not idno_set.issubset(saved_idno_set):
                saved_idno_set.add(identity_idno)               

        # -------------------
        # Process the content
        # -------------------
        # Note: this worksheet would have been skipped if it has no content.

        # Copy content worksheet information to the destination worksheet.
        #  o The identity_idno identifies the destination row index to use for each worksheet.
        #  o If the destination cell already has contents, append the additional contents using a comma delimiter.
        for dest_ws_ind, dest_ws_details in dest_ws_by_ind.items():
            dest_ws = dest_ws_details['ws']
            dest_row_idx = identity['id'][identity_idno]['dest_ws'][dest_ws_ind]['row']

            # Copy the matched identifiers values to the destination worksheet.
            for dest_de_name, dest_de_details in identifier_matching['matched'].items():
                dest_de_value = dest_de_details['value']
                dest_col_idx = ded[dest_de_name]['dest_ws'][dest_ws_ind]['column']
                dest_format = None if not 'dest_format' in ded[dest_de_name] else ded[dest_de_name]['dest_format']
                ims_util.update_dest_cell(dest_ws, dest_row_idx, dest_col_idx, dest_de_value, dest_format)

            # Copy the content columns to the destination worksheet.
            for col_idx in ws_content_cols:
                dest_de_value = ws.cell(row_idx, col_idx).value
                if dest_de_value == None:
                    continue
                dest_format = None if not 'dest_format' in ws_column_etl[col_idx] else ws_column_etl[col_idx]['dest_format']
                #Last name	First name	Nickname	ID	Email	Mobile phone	Home phone	Work phone
                # Not all content gets copied to all destination worksheets.
                if dest_ws_ind in ws_column_etl[col_idx]['dest_ws']:
                    dest_row_idx = identity['id'][identity_idno]['dest_ws'][dest_ws_ind]['row']
                    dest_col_idx = ws_column_etl[col_idx]['dest_ws'][dest_ws_ind]
                    ims_util.update_dest_cell(dest_ws, dest_row_idx, dest_col_idx, dest_de_value, dest_format)

        # Also check for and process any assembled content.
        if 'content' in assembled_de_values:

            # For each assembled destination data element:
            for dest_de_name, dest_de_details in assembled_de_values['content'].items():
                dest_de_value = dest_de_details['value']
                dest_format = None if not 'dest_format' in ded[dest_de_name] else ded[dest_de_name]['dest_format']

                # For each mapped destination worksheet:
                for dest_ws_ind in dest_de_details['dest_ws'].items():

                    # Prepare to update the worksheet.
                    dest_ws_details = dest_ws_by_ind[dest_ws_ind]
                    dest_ws = dest_ws_details['ws']

                    # Update the worksheet.
                    dest_row_idx = identity['id'][identity_idno]['dest_ws'][dest_ws_ind]['row']
                    dest_col_idx = dest_ws_details['column']
                    ims_util.update_dest_cell(dest_ws, dest_row_idx, dest_col_idx, dest_de_value, dest_format)

# Ensure that the DED WS is first, followed by each of the reports.
# wb.move_sheet(sheet, offset=0)

if not DEBUG_REPORTING_ONLY:
    wb.save(DEBUG_WORKBOOK)
wb.close()